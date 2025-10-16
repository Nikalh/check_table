# -*- coding: utf-8 -*-
import asyncio
import os
import datetime
import sys
import html

print("🚀 Starting Telegram Excel Bot on PythonAnywhere...")

# Импорты
import pandas as pd
from openpyxl import load_workbook
import requests  # Используем requests вместо aiohttp
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from dotenv import load_dotenv

# Для Render - добавляем простой веб-сервер
from aiohttp import web
import threading

# Загружаем переменные окружения
load_dotenv()
TOKEN = os.getenv("TELEGRAM_TOKEN")

if not TOKEN:
    print("❌ TELEGRAM_TOKEN not found!")
    sys.exit(1)

print("✅ Telegram token loaded successfully")

bot = Bot(token=TOKEN)
dp = Dispatcher()
scheduler = AsyncIOScheduler()
user_data = {}

# --- Экранирование текста ---
def escape_md(text):
    if not text:
        return ""
    escape_chars = r'\_*[]()~`>#+-=|{}.!'
    return ''.join(f'\\{c}' if c in escape_chars else c for c in str(text))

# --- Безопасное приведение к дате ---
def parse_date(val):
    if val is None:
        return None
    try:
        return pd.to_datetime(val).date()
    except Exception:
        return None

# --- Проверка выполненности задачи ---
def is_done(cell):
    if not cell:
        return False
    try:
        # Черный шрифт
        font_black = False
        if hasattr(cell, 'font') and cell.font:
            if hasattr(cell.font, 'color') and cell.font.color:
                if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                    font_color = str(cell.font.color.rgb).upper()
                    font_black = font_color in ["FF000000", "00000000", "000000", "FF00000000", None]
        
        # Зеленая заливка
        green_fill = False
        if hasattr(cell, 'fill') and cell.fill:
            if hasattr(cell.fill, 'fill_type') and cell.fill.fill_type is not None:
                if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                    if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                        fill_color = str(cell.fill.fgColor.rgb).upper()
                        green_fill = (
                            fill_color.startswith("FF00") or 
                            fill_color.startswith("00FF") or 
                            fill_color.startswith("0092") or
                            fill_color.startswith("FF92") or
                            "92D050" in fill_color or
                            "00FF00" in fill_color
                        )
        
        return font_black or green_fill
        
    except Exception as e:
        print(f"Error in is_done: {e}")
        return False

# --- Скачивание файла через requests ---
async def download_file(user_id):
    data = user_data.get(user_id)
    if not data:
        await bot.send_message(user_id, "❌ Данные пользователя не найдены.")
        return None
        
    file_path = f"user_{user_id}.xlsx"
    
    # Если есть ссылка - скачиваем
    if data.get("link"):
        try:
            print(f"📥 Downloading file from: {data['link']}")
            
            # Используем requests вместо aiohttp
            response = requests.get(data["link"], timeout=30)
            if response.status_code == 200:
                with open(file_path, "wb") as f:
                    f.write(response.content)
                user_data[user_id]["path"] = file_path
                print(f"✅ File downloaded successfully: {file_path}")
                return file_path
            else:
                await bot.send_message(user_id, f"❌ Ошибка скачивания: HTTP {response.status_code}")
                return None
                
        except Exception as e:
            await bot.send_message(user_id, f"❌ Ошибка скачивания файла: {e}")
            return None
    
    # Если есть локальный путь
    if data.get("path"):
        if os.path.exists(data["path"]):
            return data["path"]
        else:
            await bot.send_message(user_id, f"❌ Файл не найден по пути: {data['path']}")
            return None
    
    await bot.send_message(user_id, "❌ Файл не найден.")
    return None

# --- Проверка Excel ---
async def check_excel(user_id, notify_today=True, week_summary=False):
    print(f"🔍 Checking Excel for user {user_id}")
    
    file_path = await download_file(user_id)
    if not file_path:
        return
        
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        await bot.send_message(user_id, f"❌ Ошибка открытия файла: {e}")
        return
        
    sheet_name = "Согласование документации"
    if sheet_name not in wb.sheetnames:
        await bot.send_message(user_id, f"❌ Лист '{sheet_name}' не найден")
        return
        
    ws = wb[sheet_name]
    today = datetime.date.today()
    overdue_items = []

    headers = [cell.value for cell in ws[1]]
    col_map = {name: idx for idx, name in enumerate(headers) if name}

    days_limit = user_data[user_id].get("days", 30)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            obj_cell = row[col_map.get("Объект")] if "Объект" in col_map else None
            task_cell = row[col_map.get("Сооружение")] if "Сооружение" in col_map else None
            resp_cell = row[col_map.get("Ответственный")] if "Ответственный" in col_map else None
            subject_cell = row[col_map.get("Предмет письма")] if "Предмет письма" in col_map else None
            pg_cell = row[col_map.get("Срок от ПГ")] if "Срок от ПГ" in col_map else None
            cc_cell = row[col_map.get("Направил в ЦЦО")] if "Направил в ЦЦО" in col_map else None

            obj = obj_cell.value if obj_cell else ""
            task = task_cell.value if task_cell else ""
            resp = resp_cell.value if resp_cell else ""
            subject = subject_cell.value if subject_cell else ""

            date_pg = parse_date(pg_cell.value if pg_cell else None)
            date_cc = parse_date(cc_cell.value if cc_cell else None)

            # Проверяем только НЕ выполненные задачи
            if pg_cell and date_pg and date_pg <= today and not is_done(pg_cell):
                overdue_items.append(
                    f"📍 <b>{html.escape(str(obj))}</b>\n"
                    f"📝 {html.escape(str(task))}\n"
                    f"👤 {html.escape(str(resp))}\n"
                    f"✉ {html.escape(str(subject))}\n"
                    f"⏰ Срок от ПГ: {html.escape(str(date_pg))}"
                )
                
            if cc_cell and date_cc and (today - date_cc).days >= days_limit and not is_done(cc_cell):
                overdue_items.append(
                    f"📍 <b>{html.escape(str(obj))}</b>\n"
                    f"📝 {html.escape(str(task))}\n"
                    f"👤 {html.escape(str(resp))}\n"
                    f"✉ {html.escape(str(subject))}\n"
                    f"⏰ Направлено в ЦЦО: {html.escape(str(date_cc))} ({html.escape(str((today - date_cc).days))} дней)"
                )
                
        except Exception as e:
            print(f"Error in row {row_idx}: {e}")
            continue

    if overdue_items:
        header = f"⚠️ {len(overdue_items)-1} Просроченные задачи на сегодня {today}:\n\n" if notify_today else "📋 Сводка задач на неделю:\n\n"
        msg = header + "\n\n".join(overdue_items)
        
        chunks = [msg[i:i+4000] for i in range(0, len(msg), 4000)]
        for chunk in chunks:
            try:
                await bot.send_message(user_id, chunk, parse_mode="HTML")
            except Exception as e:
                print(f"Error sending message: {e}")
                try:
                    clean_chunk = chunk.replace('<b>', '').replace('</b>', '')
                    await bot.send_message(user_id, clean_chunk)
                except Exception as fallback_error:
                    print(f"Fallback error: {fallback_error}")
                    
        print(f"✅ Sent {len(overdue_items)} overdue items to user {user_id}")
    elif week_summary:
        await bot.send_message(user_id, "✅ Все задачи на этой неделе в срок.")
    elif notify_today:
        await bot.send_message(user_id, "✅ На сегодня просроченных задач нет.")

# --- Команды ---
@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    await message.reply(
        "🤖 Бот для отслеживания задач из Excel-таблиц\n\n"
        "Отправьте мне ссылку на Excel файл или сам файл.\n"
        "Бот проверяет задачи ежедневно в 9:00 и присылает сводку по понедельникам."
    )

@dp.message(Command("status"))
async def cmd_status(message: types.Message):
    await message.reply("✅ Бот работает на PythonAnywhere!")

@dp.message(Command("test"))
async def cmd_test(message: types.Message):
    await message.reply("🔄 Тестовая проверка...")
    await check_excel(message.from_user.id, notify_today=True)

# --- Обработка сообщений ---
@dp.message()
async def handle_message(message: types.Message):
    user_id = message.from_user.id
    
    if message.text:
        text = message.text.strip()
        
        # Проверка ссылки
        if text.lower().startswith(('http://', 'https://')):
            user_data[user_id] = {"link": text, "days": 30}
            await message.reply("✅ Ссылка принята. Проверяю таблицу...")
            await check_excel(user_id)
            return
            
        # Проверка локального файла (для отладки)
        elif os.path.exists(text) and text.lower().endswith(('.xlsx', '.xls')):
            user_data[user_id] = {"path": text, "days": 30}
            await message.reply("✅ Локальный файл принят. Проверяю таблицу...")
            await check_excel(user_id)
            return
            
        else:
            # Если это не ссылка и не существующий файл, покажем инструкцию
            await message.reply(
                "📋 Отправьте мне:\n\n"
                "• **Ссылку** на Excel файл (http://...)\n"
                "• Или **сам файл** как документ\n\n"
                "Пример ссылки: https://example.com/file.xlsx"
            )
            return
        
    elif message.document and message.document.file_name:
        file_name = message.document.file_name.lower()
        if file_name.endswith(('.xlsx', '.xls')):
            file_path = f"user_{user_id}.xlsx"
            try:
                file = await bot.get_file(message.document.file_id)
                await bot.download_file(file.file_path, destination=file_path)
                user_data[user_id] = {"path": file_path, "days": 30}
                await message.reply("✅ Файл получен. Проверяю таблицу...")
                await check_excel(user_id)
            except Exception as e:
                await message.reply(f"❌ Ошибка при сохранении файла: {e}")
        else:
            await message.reply("❌ Файл должен быть Excel (.xlsx или .xls)")
        return
        
    await message.reply("❌ Отправьте Excel файл или ссылку на него")

# --- Планировщик ---
async def daily_check():
    print("🔄 Running daily check...")
    if not user_data:
        print("❌ No users found for daily check")
        return
        
    for user_id in list(user_data.keys()):
        try:
            await check_excel(user_id, notify_today=True)
        except Exception as e:
            print(f"Error in daily check for user {user_id}: {e}")

async def weekly_summary():
    print("📋 Running weekly summary...")
    if not user_data:
        print("❌ No users found for weekly summary")
        return
        
    for user_id in list(user_data.keys()):
        try:
            await check_excel(user_id, notify_today=False, week_summary=True)
        except Exception as e:
            print(f"Error in weekly summary for user {user_id}: {e}")

# --- Простой веб-сервер для Health Checks ---
async def health_check(request):
    return web.Response(text="Bot is running!")

def run_web_server():
    app = web.Application()
    app.router.add_get('/', health_check)
    app.router.add_get('/health', health_check)
    
    # Получаем порт из переменной окружения (Render сам назначает)
    port = int(os.environ.get("PORT", 10000))
    web.run_app(app, host='0.0.0.0', port=port)


# --- Основная функция ---
""" async def main():
    print("✅ Bot initialized successfully")
    
    # Запускаем планировщик
    scheduler.add_job(daily_check, "cron", hour=6, minute=0, timezone="Europe/Moscow")  # 9:00 МСК
    scheduler.add_job(weekly_summary, "cron", day_of_week=0, hour=7, minute=0, timezone="Europe/Moscow")  # 10:00 МСК в воскресенье
    scheduler.start()
    
    print("⏰ Scheduler started: Daily at 09:00 MSK, Weekly on Sunday at 10:00 MSK")
    print("🤖 Bot is ready and polling...")
    
    # Запускаем бота
    await dp.start_polling(bot)
 """

# --- Основная функция с автоматическим перезапуском ---
async def main():
    print("✅ Bot initialized successfully")
    
    # Запускаем планировщик
    scheduler.add_job(daily_check, "cron", hour=6, minute=0, timezone="Europe/Moscow")
    scheduler.add_job(weekly_summary, "cron", day_of_week=0, hour=7, minute=0, timezone="Europe/Moscow")
    scheduler.start()
    
    print("⏰ Scheduler started")
    print("🤖 Bot is ready and polling...")
    
    # Запускаем бота с перезапуском при ошибках
    restart_count = 0
    max_restarts = 10
    
    while restart_count < max_restarts:
        try:
            await dp.start_polling(bot)
        except Exception as e:
            restart_count += 1
            print(f"❌ Bot crashed (restart {restart_count}/{max_restarts}): {e}")
            print("🔄 Restarting in 30 seconds...")
            await asyncio.sleep(30)
    
    print("❌ Max restarts reached. Bot stopped.")

if __name__ == "__main__":
   # Запускаем веб-сервер в отдельном потоке для Health Checks
    web_thread = threading.Thread(target=run_web_server, daemon=True)
    web_thread.start()
    # Для PythonAnywhere - запуск с обработкой ошибок
    asyncio.run(main())
